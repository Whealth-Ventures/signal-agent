"""Build inputs/tuning.xlsx from the project's compiled-in defaults.

This script is a one-shot "factory reset": run it to regenerate tuning.xlsx
with the values the project originally shipped with. After the first run,
tuning.xlsx is the source of truth — edit it in Excel-for-the-web, on disk,
or via GitHub's web UI. The script is checked in so the defaults are visible
in code review and recoverable from any state of the xlsx.

Usage:
    python scripts/build_default_tuning_xlsx.py             # write to inputs/tuning.xlsx
    python scripts/build_default_tuning_xlsx.py --force     # overwrite without prompt
    python scripts/build_default_tuning_xlsx.py --out X.xlsx
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUT = ROOT / "inputs" / "tuning.xlsx"


# --- The defaults ------------------------------------------------------
#
# Mirrors what was originally in src/config.py. Keep the column shape stable;
# src/tunables.py reads by header name, not column position, so re-ordering
# columns in Excel is safe.

SETTINGS: list[tuple[str, object, str]] = [
    # Budget
    ("max_perplexity_calls_per_day", 60,
     "Hard cap on Perplexity API calls per day. Enforced PER GEO RUN: the India "
     "and US runs each get their own count, so total capacity is ~2x this."),
    ("daily_budget_usd", 6.0,
     "Soft budget reference; currently used only in logging. ~2x now that India "
     "and US each run a deep sweep."),

    # Digest shape
    ("max_digest_items", 22,
     "Ceiling on total ranked stories per digest. With target_digest_min as the "
     "floor, most days land in a tight band just under this."),
    ("target_digest_min", 18,
     "Floor on total ranked stories. If normal S/A selection lands below this, "
     "the ranker backfills with the best remaining Tier-B stories so slow news "
     "days aren't thin. 0 disables the floor."),
    ("top_summary_size", 5,
     "Stories promoted into 'Today's biggest stories' at the top of the Slack post."),

    # Dedup
    ("dedup_window_days", 30,
     "How far back 'recently sent' goes for URL dedup, ranker candidate filter, and cross-day similarity."),
    ("historical_dedup_threshold", 0.80,
     "Cosine similarity above which a new story is dropped as a duplicate of one sent in the last N days."),

    # Scoring
    ("cluster_similarity_threshold", 0.85,
     "Within-day clustering threshold. Signals above this collapse into one story."),
    ("top_k_content_similarity", 5,
     "How many content corpus chunks to compare against per story. More = smoother score, slower."),
    ("summary_truncate_for_embed", 400,
     "Characters of summary text fed to the embedding model."),

    # Ranker prompt mechanics
    ("min_candidate_score", 0.0,
     "Pre-filter on relevance_score before ranker sees stories. 0.0 = no pre-filter."),
    ("one_liner_max_chars", 120,
     "Hard cap on the one-line headline. Forces newsroom-punchy output."),
    ("ranker_summary_max_chars", 220,
     "How much of each story's summary the prompt shows the ranker."),

    # Perplexity
    ("perplexity_model_fetch", "sonar-pro",
     "Model for the 30+-plan fetch sweep."),
    ("perplexity_model_rank", "sonar-reasoning-pro",
     "Model for the single ranking call."),
    ("perplexity_recency", "day",
     "Recency filter on the fetch sweep."),

    # Embeddings
    ("embedding_model", "text-embedding-3-small",
     "Embedding model for both content-corpus indexing and signal embedding. Switching requires re-indexing."),

    # HTTP
    ("http_timeout_s", 30,
     "Default timeout in seconds for fetch HTTP calls."),
    ("http_timeout_rank_s", 120,
     "Looser timeout in seconds for sonar-reasoning-pro (extended chain-of-thought)."),
    ("http_max_retries", 4,
     "Max retries on retryable HTTP errors."),
    ("url_validation_timeout_s", 10,
     "HEAD validation budget per story before posting to Slack."),

    # Schedule
    ("digest_tz", "Asia/Kolkata",
     "Intended fire timezone (the actual schedule is the systemd timer in deploy/signal-agent.timer)."),
    ("digest_hour_local", 8,
     "Intended fire hour, local time."),

    # Track B rotation
    ("track_b_plans_per_day", 40,
     "Non-priority sub-buckets sampled per day from the rotation."),
    ("track_b_rotation_days", 7,
     "Full-cycle length of the Track B rotation."),
]


# Boosters — name, weight, pattern (blank for special ones handled by name in
# scorer.compute_boosters), description.
BOOSTERS: list[tuple[str, float, str, str]] = [
    ("tier1_voice",         0.10, "",
     "Story mentions a Tier-1 voice from voices.xlsx. Matched by name, not regex."),
    ("trusted_publication", 0.08, "",
     "Story URL host matches a newsletter in voices.xlsx. Matched by host, not regex."),
    ("firm_mention",        0.08, "",
     "Story mentions a firm from the New Additions tab. Matched by name, not regex."),
    ("funding",   0.05, r"\b(raises?|series [a-d]|seed round|funding)\b",
     "Funding-round language in title/summary."),
    ("m_and_a",   0.05, r"\b(acquires?|acquisition|merges? with|m&a)\b",
     "M&A language in title/summary."),
    ("regulatory", 0.05, r"\b(fda|cdsco|ema|approved|cleared)\b",
     "Regulator language in title/summary."),
    ("product",   0.03, r"\b(launches?|unveils?|debuts?)\b",
     "Product-launch language in title/summary."),
    ("leadership", 0.03, r"\b(appoints?|named|joins|hires?)\b",
     "Leadership-move language in title/summary."),
    ("listicle",  -0.10, r"^\s*\d+\s+(best|top|essential|reasons|tips|ways)\b",
     "Penalty for listicle headlines like '10 best...'."),
    ("opinion",   -0.05, r"^\s*(opinion|perspective|column):",
     "Penalty for opinion/column headlines."),
]


# Priority buckets — sub_buckets and geos are semicolon-separated lists.
PRIORITY_BUCKETS: list[tuple[str, str, str, str]] = [
    ("venture_ipo", "Venture & IPO",
     "Venture and IPO; Capital and deal types", "India; US"),
    ("pe_strategics", "PE & Strategics",
     "Private equity and strategics", "India; US"),
    ("hospital_ma", "Hospital & Health System M&A",
     "Hospital and health system M&A", "India; US"),
    ("mso_rollups", "Physician Practice & MSO Roll-ups",
     "Physician practice and MSO roll-ups", "US"),
    ("fda_regulatory", "FDA & Regulatory",
     "FDA, regulatory and approvals", "India; US"),
    ("hot_tas", "Phase 3 / Hot Therapeutic Areas",
     "Hot therapeutic areas", "Global"),
    ("us_medicare", "US Medicare",
     "US Medicare policy and reform", "US"),
    ("ai_healthcare", "AI in Healthcare",
     "Generative AI and LLMs in healthcare; AI scribes and ambient documentation; Revenue cycle and administrative AI",
     "Global"),
]


# Source tiers — ordered list of hosts. Earlier = preferred when picking the
# canonical URL among cluster duplicates.
SOURCE_TIERS: list[str] = [
    # US — finance / business / general
    "bloomberg.com", "reuters.com", "wsj.com", "nytimes.com", "ft.com",
    "axios.com", "cnbc.com", "forbes.com",
    # US — healthcare-specific
    "statnews.com", "endpts.com", "biopharmadive.com", "modernhealthcare.com",
    "fiercehealthcare.com", "fiercebiotech.com", "fiercepharma.com",
    "healthcaredive.com", "medcitynews.com", "beckershospitalreview.com",
    "healthcareitnews.com", "healthaffairs.org",
    # India — finance / business / general
    "economictimes.indiatimes.com", "livemint.com", "business-standard.com",
    "moneycontrol.com", "thehindubusinessline.com", "financialexpress.com",
    "thehindu.com", "indianexpress.com",
    # India — healthcare-specific
    "health.economictimes.indiatimes.com", "biospectrumindia.com",
    "expresshealthcare.in", "pharmabiz.com",
    # India — tech / startup news
    "inc42.com", "yourstory.com", "entrackr.com", "the-ken.com",
]


# --- Styling -----------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="E2E8F0")
_HEADER_FONT = Font(bold=True)


def _write_sheet(ws, headers: list[str], rows: list[tuple], widths: list[int]) -> None:
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="left")
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def build(path: Path) -> None:
    wb = Workbook()

    # Sheet 1: Settings
    ws = wb.active
    ws.title = "Settings"
    _write_sheet(
        ws,
        headers=["name", "value", "description"],
        rows=SETTINGS,
        widths=[35, 25, 90],
    )

    # Sheet 2: Boosters
    ws2 = wb.create_sheet("Boosters")
    _write_sheet(
        ws2,
        headers=["name", "weight", "pattern_regex", "description"],
        rows=BOOSTERS,
        widths=[24, 10, 55, 80],
    )

    # Sheet 3: Priority Buckets
    ws3 = wb.create_sheet("Priority Buckets")
    _write_sheet(
        ws3,
        headers=["key", "display", "sub_buckets", "geos"],
        rows=PRIORITY_BUCKETS,
        widths=[20, 38, 90, 16],
    )

    # Sheet 4: Source Tiers
    ws4 = wb.create_sheet("Source Tiers")
    _write_sheet(
        ws4,
        headers=["host"],
        rows=[(h,) for h in SOURCE_TIERS],
        widths=[42],
    )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--out", type=Path, default=DEFAULT_OUT,
                   help=f"Output path (default: {DEFAULT_OUT.relative_to(ROOT)})")
    p.add_argument("--force", action="store_true",
                   help="Overwrite existing file without prompt.")
    args = p.parse_args()

    out: Path = args.out
    if out.exists() and not args.force:
        ans = input(f"{out} exists. Overwrite? [y/N] ").strip().lower()
        if ans != "y":
            print("Aborted.")
            return
    build(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
