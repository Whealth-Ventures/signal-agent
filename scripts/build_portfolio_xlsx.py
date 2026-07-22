"""Build inputs/portfolio.xlsx from the compiled-in W Health / 2070 Health list.

One-shot seeder for the Sector Agent's portfolio input. After the first run,
portfolio.xlsx is the source of truth — edit it on disk or via the admin UI's
Portfolio editor (which rewrites the same data region in place).

Layout (kept deliberately flat so the loaders are trivial — mirrors
keywords.xlsx, NOT the banner/offset style of voices.xlsx):
    Sheet "Portfolio"
    Row 1 = header:  Company | Sector | What they do | Geo | Website
    Row 2+ = data

`Geo` is one of India / US / Global (Global = multi-geo / cross-border), used
by the sector query planner to frame "…in that geography" alongside global
developments. Values are best-effort seeds — edit them in the admin UI.

Usage:
    python scripts/build_portfolio_xlsx.py            # write inputs/portfolio.xlsx
    python scripts/build_portfolio_xlsx.py --force    # overwrite without prompt
    python scripts/build_portfolio_xlsx.py --out X.xlsx
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_OUT = ROOT / "inputs" / "portfolio.xlsx"

HEADER = ["Company", "Sector", "What they do", "Geo", "Website"]

# Company | Sector | What they do | Geo | Website
# Seeded from whealthventures.com/portfolio, enriched with per-company research
# (2026-07). Descriptions are information-dense on purpose — they feed the impact
# ranker's prompt. Fuller context (competitors, what moves each company) lives in
# inputs/portfolio_context.md. Website left blank (a wrong URL is worse than none);
# fill via the admin Portfolio page.
PORTFOLIO: list[tuple[str, str, str, str, str]] = [
    ("BeatO", "Digital diabetes management",
     "App + smartphone-glucometer platform selling test strips/CGMs and subscription "
     "coaching (diet, doctor, GLP-1 weight programs) to Indian diabetics and pre-diabetics",
     "India", ""),
    ("Mylo", "Mother & baby (community + D2C)",
     "Pregnancy/parenting app and mom community monetizing via its own D2C brands "
     "(baby care, mother wellness, Ayurveda) for expecting and new mothers", "India", ""),
    ("Good Health Company", "Digital wellness clinic (D2C)",
     "Digital clinic (Mars for men, Saturn for women): free online consults + prescribed "
     "courses and D2C products across sexual health, hair, skin, and weight loss", "India", ""),
    ("Elevate Now", "Medical (GLP-1) weight loss",
     "Doctor-led medical weight-loss program combining prescription GLP-1 pharmacotherapy "
     "with coaching, nutrition, and a non-surgical gastric-balloon option", "India", ""),
    ("BabyMD", "Pediatric primary care (clinics)",
     "Tech-enabled pediatric clinic network (vaccinations, consults, developmental "
     "assessments, 24x7 AI-assisted doctor access) for urban parents; recently rebranded "
     "Hoola Health", "India", ""),
    ("Nivaan Care", "Pain management clinics",
     "Single-specialty multidisciplinary chronic-pain clinic chain (pain physicians + "
     "physio + minimally invasive daycare interventional procedures) across North & West India",
     "India", ""),
    ("2070Health", "Healthcare venture studio",
     "W Health's healthcare venture studio building companies from scratch in-house "
     "(e.g. Elevate Now, Nivaan, Reveal, Everhope)", "India", ""),
    ("Everhope Oncology", "Oncology daycare centers",
     "Chain of dedicated cancer daycare/infusion centers (chemo, targeted therapy, "
     "precision-oncology, diagnostics, supportive care) in a non-hospital setting", "India", ""),
    ("Stealth B2B Services Co.", "B2B services (stealth)",
     "Stealth-stage B2B services provider across the India/US corridor; no verifiable "
     "public product info yet", "Global", ""),
    ("Wysa", "AI mental-health / digital therapeutics",
     "Clinically-validated AI-guided mental-health app (CBT + coaching) sold D2C and to "
     "employers, health plans, and health systems (e.g. NHS); acquired Kins (PT) and April Health",
     "Global", ""),
    ("Jasper Health", "Digital oncology / cancer navigation",
     "Virtual 1-on-1 cancer support (oncology-trained guides + planning platform) sold to "
     "health plans and self-insured employers, incl. a Medicare navigation offering", "US", ""),
    ("Kins", "Hybrid MSK / physical therapy",
     "Hybrid PT: 45-55 min 1-on-1 at-home or virtual sessions by licensed PTs with "
     "between-visit digital engagement; accepts most insurance/Medicare; now part of Wysa "
     "(acquired Sept 2025)", "US", ""),
    ("Violet Health", "Health equity / clinician upskilling",
     "SaaS measuring clinicians' cultural competence and delivering upskilling + "
     "credentialing so provider orgs and payers deliver identity-centered care for BIPOC, "
     "LGBTQ+, and underserved patients", "US", ""),
    ("Reveal HealthTech", "Healthcare AI & engineering services",
     "Data/AI/engineering services firm (tools: BioCanvas trial recruitment, Prism AI ops) "
     "building bespoke AI for US healthcare/life-sciences with India-based delivery", "Global", ""),
    ("Ryse Health", "Value-based diabetes care",
     "Hybrid in-person + virtual specialty clinics for uncontrolled Type 2 diabetes under "
     "payer value-based/risk contracts (CGM + app), paid on outcomes like A1c reduction", "US", ""),
    ("Everbright Health", "Behavioral health enablement",
     "Tech-enabled platform letting mental-health practices launch advanced interventions "
     "(TMS, SPRAVATO/esketamine) via clinical infra, staffing, AI patient ID, and "
     "prior-auth/billing management", "US", ""),
]


def build(out: Path, force: bool) -> None:
    if out.exists() and not force:
        resp = input(f"{out} exists. Overwrite? [y/N] ").strip().lower()
        if resp != "y":
            print("Aborted.")
            return

    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    for col, name in enumerate(HEADER, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center")

    for r, row in enumerate(PORTFOLIO, start=2):
        for col, val in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=val)

    widths = [26, 26, 60, 10, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote {len(PORTFOLIO)} companies to {out}")


def main() -> None:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--out", type=Path, default=DEFAULT_OUT)
    p.add_argument("--force", action="store_true")
    args = p.parse_args()
    build(args.out, args.force)


if __name__ == "__main__":
    main()
