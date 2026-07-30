"""Bootstrap inputs/sectors.xlsx for the bi-weekly sector digest.

Writes three tabs — Sectors, Sector Keywords, Sector Watchlist — with starter
content for the 7 launch sectors. Keywords are tagged Global (the per-plan
prompt's geo label does the regional scoping); edit/extend freely in Excel or
the admin UI. Re-running overwrites the file.

    python scripts/build_sectors_xlsx.py
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))
import config  # noqa: E402


# key, display, portco, primary_geo, secondary_geo, secondary_max, target, notes
SECTORS: list[dict] = [
    dict(
        key="mental_health", display="Mental & Behavioral Health",
        portco="Everbright", primary_geo="US", secondary_geo="", secondary_max=0,
        target=15,
        notes="US focus with emphasis on interventional psychiatry (TMS, ketamine/esketamine, ECT).",
        keywords=[
            "mental health", "behavioral health", "interventional psychiatry",
            "TMS", "transcranial magnetic stimulation", "ketamine", "esketamine",
            "Spravato", "treatment-resistant depression", "psychedelic therapy",
            "ECT", "digital mental health", "teletherapy", "substance use disorder",
            "ABA therapy", "psychiatry", "depression", "anxiety",
        ],
        watchlist=[
            "Talkiatry", "Rula", "Grow Therapy", "Headway", "SonderMind",
            "Osmind", "Spring Health", "Lyra Health", "Talkspace", "Brightside",
            "Two Chairs", "Compass Pathways", "atai Life Sciences", "Greenbrook TMS",
        ],
    ),
    dict(
        key="pediatrics", display="Pediatrics", portco="Hoola Health",
        primary_geo="India", secondary_geo="US", secondary_max=3, target=15,
        notes="India pediatrics + a few biggest US pediatrics stories.",
        keywords=[
            "pediatrics", "pediatric care", "child health", "neonatal", "NICU",
            "children's hospital", "pediatric telehealth", "childhood illness",
            "pediatric surgery", "adolescent health", "pediatric clinic",
            "child vaccination", "newborn care", "pediatric hospital",
        ],
        watchlist=[
            "Rainbow Children's Hospital", "Cloudnine", "Ovum Hospitals",
            "Motherhood Hospitals", "Ankura Hospital", "Summer Health",
            "Cradlewise", "Boston Children's Hospital",
            "Children's Hospital of Philadelphia",
        ],
    ),
    dict(
        key="diabetes", display="Diabetes", portco="Beato",
        primary_geo="India", secondary_geo="", secondary_max=0, target=15,
        notes="",
        keywords=[
            "diabetes", "type 2 diabetes", "type 1 diabetes", "diabetes management",
            "CGM", "continuous glucose monitor", "insulin", "GLP-1", "diabetes care",
            "blood glucose", "HbA1c", "prediabetes", "diabetic retinopathy",
            "metabolic health", "digital diabetes", "diabetes reversal",
        ],
        watchlist=[
            "BeatO", "Fitterfly", "Twin Health", "Sugar.fit", "Phable",
            "Wellthy Therapeutics", "Novo Nordisk", "Eli Lilly", "Abbott",
            "Dexcom",
        ],
    ),
    dict(
        key="pain_management", display="Pain Management", portco="Nivaan Care",
        primary_geo="India", secondary_geo="", secondary_max=0, target=15,
        notes="",
        keywords=[
            "pain management", "chronic pain", "interventional pain",
            "radiofrequency ablation", "nerve block", "spinal cord stimulation",
            "back pain", "musculoskeletal pain", "pain clinic", "palliative care",
            "physiotherapy", "neuropathic pain", "pain relief", "orthopedic pain",
        ],
        watchlist=[
            "Nivaan Care", "QI Spine Clinic", "Sukino Healthcare", "Physiotattva",
            "Portea Medical", "Boston Scientific", "Medtronic",
        ],
    ),
    dict(
        key="parenting", display="Parenting", portco="Mylo",
        primary_geo="India", secondary_geo="", secondary_max=0, target=15,
        notes="Consumer/health blend — watch sector topicality.",
        keywords=[
            "parenting", "maternity", "pregnancy", "prenatal", "postnatal",
            "new parents", "baby care", "infant care", "child development",
            "breastfeeding", "fertility", "motherhood", "parenting app",
            "maternal health", "newborn",
        ],
        watchlist=[
            "Mylo", "FirstCry", "BabyChakra", "The Moms Co", "Tinystep",
            "Proactive For Her", "Kindlife", "Mamaearth",
        ],
    ),
    dict(
        key="weight_loss", display="Doctor-led Medical Weight Loss",
        portco="ElevateNow", primary_geo="India", secondary_geo="",
        secondary_max=0, target=15, notes="GLP-1 / obesity clinics.",
        keywords=[
            "medical weight loss", "weight loss", "obesity", "GLP-1",
            "semaglutide", "Wegovy", "Ozempic", "tirzepatide", "Mounjaro",
            "weight management", "bariatric", "metabolic health",
            "anti-obesity medication", "obesity clinic", "weight loss drug",
        ],
        watchlist=[
            "ElevateNow", "Fitterfly", "Fitelo", "HealthifyMe", "Ultrahuman",
            "Novo Nordisk", "Eli Lilly", "Zydus", "Allurion",
        ],
    ),
    dict(
        key="oncology", display="Oncology", portco="Everhope",
        primary_geo="India", secondary_geo="US", secondary_max=3, target=15,
        notes="India oncology + a few biggest US oncology stories.",
        keywords=[
            "oncology", "cancer", "cancer care", "cancer treatment", "chemotherapy",
            "immunotherapy", "radiation oncology", "cancer hospital", "tumor",
            "cancer diagnostics", "precision oncology", "cancer drug",
            "CAR-T", "oncology clinic", "cancer screening",
        ],
        watchlist=[
            "Everhope", "HealthCare Global", "HCG", "Karkinos Healthcare",
            "Cytecare", "American Oncology Institute", "Apollo Cancer Centres",
            "Onco.com", "Thyme Care", "Flatiron Health",
        ],
    ),
]


# Per-sector authors / newsletters / reports to name in the sweep so Perplexity
# surfaces their recent coverage. (type, name, url). Starter list from research —
# review and edit freely. India sectors lean on general India health/startup
# outlets where a dedicated sector newsletter doesn't exist yet.
SECTOR_SOURCES: dict[str, list[tuple[str, str, str]]] = {
    "mental_health": [
        ("author", "The Hemingway Report (Steve Duke)", "https://thehemingwayreport.beehiiv.com"),
        ("newsletter", "Behavioral Health Business", "https://bhbusiness.com"),
        ("newsletter", "Behavioral Health Tech", "https://www.behavioralhealthtech.com"),
        ("newsletter", "STAT News — Mental Health", "https://www.statnews.com/topic/mental-health/"),
        ("newsletter", "Behavioral Health News", "https://behavioralhealthnews.org"),
    ],
    "pediatrics": [
        ("report", "Children's Hospital Association — Landscape Report", "https://www.childrenshospitals.org"),
        ("newsletter", "Fierce Healthcare", "https://www.fiercehealthcare.com"),
        ("newsletter", "STAT News", "https://www.statnews.com"),
        ("newsletter", "ETHealthworld (Economic Times)", "https://health.economictimes.indiatimes.com"),
        ("newsletter", "Inc42 — Healthtech", "https://inc42.com"),
    ],
    "diabetes": [
        ("newsletter", "Drug Delivery Business", "https://www.drugdeliverybusiness.com"),
        ("report", "Grand View Research — Digital Diabetes Management", "https://www.grandviewresearch.com/industry-analysis/digital-diabetes-management-market"),
        ("report", "American Diabetes Association — Scientific Sessions", "https://diabetes.org"),
        ("newsletter", "ETHealthworld (Economic Times)", "https://health.economictimes.indiatimes.com"),
        ("newsletter", "Inc42 — Healthtech", "https://inc42.com"),
    ],
    "pain_management": [
        ("newsletter", "ETHealthworld (Economic Times)", "https://health.economictimes.indiatimes.com"),
        ("newsletter", "The Ken", "https://the-ken.com"),
        ("newsletter", "Inc42 — Healthtech", "https://inc42.com"),
        ("report", "MarketsandMarkets — Pain Management Devices", "https://www.marketsandmarkets.com"),
    ],
    "parenting": [
        ("report", "IMARC — India Mother & Child Healthcare", "https://www.imarcgroup.com/india-mother-and-child-healthcare-market"),
        ("newsletter", "YourStory", "https://yourstory.com"),
        ("newsletter", "Entrackr", "https://entrackr.com"),
        ("newsletter", "Inc42", "https://inc42.com"),
        ("report", "Tracxn — Mom & Baby Care India", "https://tracxn.com"),
    ],
    "weight_loss": [
        ("report", "IQVIA — Obesity Outlook", "https://www.iqvia.com"),
        ("newsletter", "Endpoints News", "https://endpts.com"),
        ("newsletter", "Fierce Pharma", "https://www.fiercepharma.com"),
        ("newsletter", "ETHealthworld (Economic Times)", "https://health.economictimes.indiatimes.com"),
        ("newsletter", "Inc42 — Healthtech", "https://inc42.com"),
    ],
    "oncology": [
        ("newsletter", "Fierce Oncology (Fierce Biotech)", "https://www.fiercebiotech.com"),
        ("newsletter", "STAT News — Oncology", "https://www.statnews.com/topic/oncology/"),
        ("newsletter", "Endpoints News", "https://endpts.com"),
        ("report", "Stout — Oncology Industry Outlook", "https://www.stout.com"),
        ("newsletter", "ETHealthworld (Economic Times)", "https://health.economictimes.indiatimes.com"),
    ],
}


def main() -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "Sectors"
    ws.append([
        "key", "display", "portco", "primary_geo", "secondary_geo",
        "secondary_max_stories", "target_story_count", "notes",
    ])
    for s in SECTORS:
        ws.append([
            s["key"], s["display"], s["portco"], s["primary_geo"],
            s["secondary_geo"], s["secondary_max"], s["target"], s["notes"],
        ])

    kw = wb.create_sheet("Sector Keywords")
    kw.append(["sector", "keyword", "geo"])
    for s in SECTORS:
        for term in s["keywords"]:
            kw.append([s["key"], term, "Global"])

    wl = wb.create_sheet("Sector Watchlist")
    wl.append(["sector", "company"])
    for s in SECTORS:
        for company in s["watchlist"]:
            wl.append([s["key"], company])

    src = wb.create_sheet("Sector Sources")
    src.append(["sector", "type", "name", "url", "rss_url"])
    for s in SECTORS:
        for type_, name, url in SECTOR_SOURCES.get(s["key"], []):
            src.append([s["key"], type_, name, url, ""])

    config.SECTORS_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(config.SECTORS_XLSX)
    n_kw = sum(len(s["keywords"]) for s in SECTORS)
    n_wl = sum(len(s["watchlist"]) for s in SECTORS)
    n_src = sum(len(v) for v in SECTOR_SOURCES.values())
    print(f"Wrote {config.SECTORS_XLSX}")
    print(f"  {len(SECTORS)} sectors, {n_kw} keywords, {n_wl} watchlist, {n_src} sources")


if __name__ == "__main__":
    main()
