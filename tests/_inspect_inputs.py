"""One-off inspection of inputs/. Not part of runtime — kept under tests/ for reproducibility."""
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent


def excel_summary(path: Path) -> None:
    print(f"\n=== {path.name} ===")
    wb = load_workbook(path, read_only=True, data_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print(f"  [{name}] empty")
            continue
        header = rows[0]
        non_empty = [r for r in rows[1:] if any(c is not None and str(c).strip() for c in r)]
        print(f"  [{name}] cols={len(header)} header_row_count={len(rows)} data_rows(non_empty)={len(non_empty)}")
        print(f"    headers: {[str(h).strip() if h else '' for h in header]}")
        # 2 sample data rows
        for r in non_empty[:2]:
            cleaned = [str(c).strip() if c is not None else '' for c in r]
            print(f"    sample: {cleaned}")


def content_summary(root: Path) -> None:
    print("\n=== inputs/content/ ===")
    for sub in sorted(p for p in root.iterdir() if p.is_dir()):
        files = [f for f in sub.rglob('*') if f.is_file() and not f.name.startswith('.')]
        ext_counts: dict[str, int] = {}
        for f in files:
            ext_counts[f.suffix.lower() or '<noext>'] = ext_counts.get(f.suffix.lower() or '<noext>', 0) + 1
        print(f"  {sub.name}/ files={len(files)} ext={ext_counts}")


if __name__ == '__main__':
    excel_summary(ROOT / 'inputs' / 'keywords.xlsx')
    excel_summary(ROOT / 'inputs' / 'voices.xlsx')
    excel_summary(ROOT / 'inputs' / 'tuning.xlsx')
    content_summary(ROOT / 'inputs' / 'content')
