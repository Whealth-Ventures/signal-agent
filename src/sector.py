"""Sector Agent library — the weekly, portfolio-company-grouped digest.

This is the third agent (alongside the India and US daily digests). It reuses
the daily pipeline's transport layers wholesale — Perplexity fetch
(`main.fetch_perplexity_async`), dedup/clustering (`scorer.run_scoring`),
SQLite storage, the ranking-vendor selection (`ranker._build_ranker_client`),
and the Slack block/transport helpers — but swaps the two pieces whose SHAPE
genuinely differs from the daily digest:

  1. The query planner: one plan PER PORTFOLIO COMPANY (`build_sector_plans`),
     with `bucket=<company name>` so the company tag rides through
     parse → dedup (scorer.pick_bucket) into `Story.bucket`.
  2. The ranker: a single "material impact" call (`rank_impact`) that assigns
     each candidate story to the company it most affects, an impact direction,
     and a materiality tier — instead of the daily healthcare-magnitude rubric.

Output is grouped by company (`build_sector_blocks`) and posted to the sector
channel. The orchestrator lives in `sector_main.py`; it runs against
`config.SECTOR_DB_PATH` so sector stories never touch the daily candidate pool.
"""
from __future__ import annotations

import asyncio
import json
import time
from dataclasses import dataclass, field
from datetime import datetime, timezone
from functools import lru_cache
from pathlib import Path
from typing import Literal

import httpx
from openpyxl import load_workbook

import config
import ranker
import slack_client
from models import Story
from query_planner import PlanGeo, QueryPlan, _s, _slug
from slack_client import SlackResult

Direction = Literal["positive", "negative", "mixed"]
Materiality = Literal["high", "medium"]

_VALID_DIRECTIONS: frozenset[str] = frozenset({"positive", "negative", "mixed"})
_VALID_MATERIALITY: frozenset[str] = frozenset({"high", "medium"})

# Bullet prefixes by impact direction.
_DIR_MARK: dict[str, str] = {
    "positive": "🟢 ↑",
    "negative": "🔴 ↓",
    "mixed":    "🟡 ↔",
}

ONE_LINER_MAX_CHARS = config.ONE_LINER_MAX_CHARS


# --- Portfolio input ----------------------------------------------------

@dataclass(frozen=True)
class Company:
    name: str
    sector: str
    business: str
    geo: str        # "India" | "US" | "Global" (free text; coerced for plans)
    website: str = ""


@lru_cache(maxsize=1)
def load_portfolio() -> list[Company]:
    """Read inputs/portfolio.xlsx (`Portfolio` tab). Flat layout: row 1 header,
    data from row 2. Columns A-E: Company, Sector, What they do, Geo, Website.
    A row counts only if it has a company name."""
    wb = load_workbook(config.PORTFOLIO_XLSX, read_only=True, data_only=True)
    ws = wb["Portfolio"]
    out: list[Company] = []
    for r in ws.iter_rows(values_only=True, min_row=2):
        if not r:
            continue
        cells = list(r) + [None] * max(0, 5 - len(r))
        name = _s(cells[0])
        if not name:
            continue
        out.append(Company(
            name=name,
            sector=_s(cells[1]),
            business=_s(cells[2]),
            geo=_s(cells[3]) or "Global",
            website=_s(cells[4]),
        ))
    return out


def _plan_geo(geo: str) -> PlanGeo:
    g = geo.strip().lower()
    if g == "india":
        return "India"
    if g in ("us", "usa", "united states"):
        return "US"
    return "Global"


def _geo_phrase(geo: PlanGeo) -> str:
    return {"India": "India", "US": "the United States"}.get(geo, "its home market")


# --- Query planner: one plan per company --------------------------------

_SECTOR_PROMPT_TEMPLATE = (
    "Developments from the LAST 7 DAYS that could MATERIALLY affect the business "
    "of {company} — a healthcare company in {sector}: {business}. "
    "Its main operating geography is {geo_phrase}.\n"
    "Surface, in {geo_phrase} or globally:\n"
    "  (a) sector/industry shifts in its category (demand, pricing, standards of "
    "care, category funding climate, large entrants/exits);\n"
    "  (b) regulatory, reimbursement, or policy changes affecting it;\n"
    "  (c) macro shifts changing its market size, costs, or capital access;\n"
    "  (d) moves by DIRECT COMPETITORS — funding, launches, M&A, pricing, exits.\n"
    "Do NOT return {company}'s OWN funding, product, hiring, or PR news — only the "
    "world around it. Skip listicles, opinion pieces, and thought-leadership.\n"
    "Return ONLY a JSON object (no markdown fences, no preamble):\n"
    "{{\n"
    '  "stories": [\n'
    '    {{"title": "headline", "url": "source URL", '
    '"published": "ISO 8601 datetime or null", "summary": "2-sentence summary"}}\n'
    "  ]\n"
    "}}"
)


def build_sector_plans(companies: list[Company] | None = None) -> list[QueryPlan]:
    """One QueryPlan per portfolio company. `bucket=company.name` is the load-
    bearing bit: parse_perplexity_response stamps it into signal.raw['bucket'],
    and scorer.pick_bucket carries it into Story.bucket through dedup — so each
    surviving story knows which company's sweep surfaced it."""
    companies = companies if companies is not None else load_portfolio()
    plans: list[QueryPlan] = []
    for c in companies:
        geo = _plan_geo(c.geo)
        prompt = _SECTOR_PROMPT_TEMPLATE.format(
            company=c.name,
            sector=c.sector or "healthcare",
            business=c.business or "(no description)",
            geo_phrase=_geo_phrase(geo),
        )
        plans.append(QueryPlan(
            id=f"sector__{_slug(c.name)}",
            geography=geo,
            bucket=c.name,
            sub_buckets=(),
            keyword_sample=(),
            keyword_count_total=0,
            prompt_text=prompt,
            track="sector",
            priority_bucket=None,
        ))
    return plans


# --- Impact ranking -----------------------------------------------------

@dataclass(frozen=True)
class ImpactItem:
    story: Story
    company: str
    materiality: Materiality
    direction: Direction
    one_liner: str


@dataclass(frozen=True)
class SectorResult:
    grouped: dict[str, list[ImpactItem]]   # company name → items (portfolio order)
    candidates_count: int
    used_fallback: bool
    cost_usd: float
    elapsed_seconds: float
    flat: tuple[ImpactItem, ...] = field(default_factory=tuple)


_MATERIALITY_RANK: dict[str, int] = {"high": 0, "medium": 1}


def _trim(s: str, n: int) -> str:
    s = (s or "").replace("\n", " ").strip()
    return s if len(s) <= n else s[: n - 1].rstrip() + "…"


@lru_cache(maxsize=1)
def load_portfolio_context() -> str:
    """The optional deeper knowledge base (competitors + what moves each company),
    appended to the impact prompt. Empty string when the file is absent."""
    try:
        return config.PORTFOLIO_CONTEXT_MD.read_text(encoding="utf-8").strip()
    except (FileNotFoundError, OSError):
        return ""


def build_impact_prompt(
    stories: list[Story], companies: list[Company], *, context: str = "",
) -> str:
    lines: list[str] = [
        "You are triaging external news for its material impact on a healthcare "
        "VC's portfolio companies.",
        "",
        config.SECTOR_IMPACT_RUBRIC,
        "",
        "Portfolio companies:",
    ]
    for c in companies:
        lines.append(f"  - {c.name} ({c.geo}) — {c.sector}: {c.business}")
    if context:
        lines += [
            "",
            "Reference — each company's competitors and what materially moves it "
            "(use this to judge relevance, competitor moves, and impact direction):",
            context,
        ]
    lines += [
        "",
        "For EACH story that is MATERIAL (high or medium) to exactly one company, "
        "return an object with:",
        "  - `story_id`: the id from the candidate list below",
        "  - `company`: the EXACT company name from the list above",
        f"  - `materiality`: \"high\" or \"medium\"",
        "  - `direction`: \"positive\", \"negative\", or \"mixed\" (impact on THAT "
        "company's business)",
        f"  - `one_liner`: one sentence (max ~{ONE_LINER_MAX_CHARS} chars) — lead "
        "with the external event, then the 'so what' for the company",
        "DROP low-materiality, off-topic, or a company's-own-PR stories: simply "
        "omit them from the output.",
        "Return ONLY a JSON object (no markdown fences):",
        '{ "items": [ {"story_id": "...", "company": "...", "materiality": '
        '"high", "direction": "negative", "one_liner": "..."} ] }',
        "",
        f"Candidate stories ({len(stories)}). `surfaced_for` is the company whose "
        "search returned it — a hint, not binding:",
    ]
    for st in stories:
        lines.append(f"[id={st.id}  surfaced_for={st.bucket or '-'}]")
        lines.append(f"  title: {_trim(st.canonical_title, 240)}")
        lines.append(f"  summary: {_trim(st.canonical_summary, config.RANKER_SUMMARY_MAX_CHARS)}")
        lines.append(f"  url: {st.canonical_url}")
    return "\n".join(lines)


def _parse_impact(
    response_text: str,
    stories_by_id: dict[str, Story],
    company_by_lower: dict[str, str],
) -> list[ImpactItem]:
    """Parse the LLM JSON into ImpactItems. Silently skips entries with an
    unknown story_id/company or an invalid tier — the fallback path (in
    rank_impact) covers a wholesale parse failure."""
    parsed = ranker._extract_json(response_text)
    if not (parsed and isinstance(parsed.get("items"), list)):
        return []
    out: list[ImpactItem] = []
    seen: set[str] = set()
    for e in parsed["items"]:
        if not isinstance(e, dict):
            continue
        sid = e.get("story_id")
        if not isinstance(sid, str) or sid not in stories_by_id or sid in seen:
            continue
        company = company_by_lower.get(str(e.get("company", "")).strip().lower())
        if not company:
            continue
        materiality = str(e.get("materiality", "")).strip().lower()
        if materiality not in _VALID_MATERIALITY:
            continue
        direction = str(e.get("direction", "")).strip().lower()
        if direction not in _VALID_DIRECTIONS:
            direction = "mixed"
        one_liner = _trim(str(e.get("one_liner") or ""), ONE_LINER_MAX_CHARS) \
            or _trim(stories_by_id[sid].canonical_title, ONE_LINER_MAX_CHARS)
        seen.add(sid)
        out.append(ImpactItem(
            story=stories_by_id[sid], company=company,
            materiality=materiality, direction=direction, one_liner=one_liner,
        ))
    return out


def _group(items: list[ImpactItem], companies: list[Company]) -> dict[str, list[ImpactItem]]:
    """Group items by company in portfolio order; drop empty companies. Within a
    company: high materiality first, then most-recent first."""
    by_company: dict[str, list[ImpactItem]] = {}
    for it in items:
        by_company.setdefault(it.company, []).append(it)
    grouped: dict[str, list[ImpactItem]] = {}
    for c in companies:
        bucket = by_company.get(c.name)
        if not bucket:
            continue
        bucket.sort(key=lambda x: (
            _MATERIALITY_RANK[x.materiality], -x.story.published_at.timestamp(),
        ))
        grouped[c.name] = bucket
    return grouped


def rank_impact(
    stories: list[Story],
    companies: list[Company],
    *,
    client: ranker._RankerClient | None = None,
) -> SectorResult:
    """Single LLM call: assign each candidate story to the company it most
    materially affects, with a direction + materiality tier. Falls back to
    grouping by the surfacing company (Story.bucket) if the call/parse fails, so
    the digest always ships."""
    start = time.monotonic()
    if not stories:
        return SectorResult({}, 0, False, 0.0, round(time.monotonic() - start, 3))

    stories_by_id = {st.id: st for st in stories}
    company_by_lower = {c.name.lower(): c.name for c in companies}

    rank_model = config.PERPLEXITY_MODEL_RANK
    if client is None:
        client, rank_model = ranker._build_ranker_client()

    response_text = ""
    cost = 0.0
    call_error: str | None = None
    try:
        resp = client.complete(
            build_impact_prompt(stories, companies, context=load_portfolio_context()),
            model=rank_model,
            query_id="sector-rank",
            system=config.SECTOR_SYSTEM_PROMPT,
            timeout=config.HTTP_TIMEOUT_RANK_S,
        )
        response_text = resp.text
        cost = resp.estimated_cost_usd
    except Exception as e:
        call_error = f"{type(e).__name__}: {e}"

    items = _parse_impact(response_text, stories_by_id, company_by_lower)
    used_fallback = bool(call_error) or not items
    if used_fallback:
        # Degrade to: keep every candidate whose surfacing company is known,
        # medium/mixed, one-liner = title. Never ship an empty digest on a
        # transient ranker failure.
        items = [
            ImpactItem(
                story=st,
                company=company_by_lower[(st.bucket or "").lower()],
                materiality="medium",
                direction="mixed",
                one_liner=_trim(st.canonical_title, ONE_LINER_MAX_CHARS),
            )
            for st in stories
            if (st.bucket or "").lower() in company_by_lower
        ]

    grouped = _group(items, companies)
    flat = tuple(it for c in companies for it in grouped.get(c.name, []))
    elapsed = round(time.monotonic() - start, 3)
    _log({
        "step": "sector_rank",
        "candidates": len(stories),
        "kept": len(flat),
        "companies_with_items": list(grouped.keys()),
        "used_fallback": used_fallback,
        "call_error": call_error,
        "cost_usd": cost,
        "elapsed_seconds": elapsed,
        "response_text": response_text[:2000],
    })
    return SectorResult(
        grouped=grouped,
        candidates_count=len(stories),
        used_fallback=used_fallback,
        cost_usd=cost,
        elapsed_seconds=elapsed,
        flat=flat,
    )


# --- Slack formatting (grouped by company) ------------------------------

def _bullet(it: ImpactItem) -> str:
    mark = _DIR_MARK.get(it.direction, "•")
    text = slack_client._escape_mrkdwn(it.one_liner or it.story.canonical_title)
    return f"{mark} {text} (<{it.story.canonical_url}|Link>)"


def build_sector_blocks(
    result: SectorResult, *, digest_date: str, test_mode: bool = False,
) -> list[dict]:
    """Block Kit payload grouped by company. Reuses slack_client's section
    helpers (escaping, char-limit splitting, dividers) so we don't reinvent the
    Block Kit plumbing. ponytail: reuse the daily formatter's primitives."""
    total = len(result.flat)
    plural = "story" if total == 1 else "stories"
    prefix = "[TEST] " if test_mode else ""
    blocks: list[dict] = [
        slack_client._section(
            f"*{prefix}Sector Signal — portfolio watch — "
            f"{slack_client._escape_mrkdwn(digest_date)}*"
            f"  ·  {total} {plural} across {len(result.grouped)} companies"
        ),
    ]
    if total == 0:
        blocks.append(slack_client._section(
            "_No material sector developments for the portfolio this week. The "
            "agent ran without errors but had no qualifying signals._"
        ))
        return blocks

    slack_client._append_divider(blocks)
    for company, items in result.grouped.items():
        header = f"*{slack_client._escape_mrkdwn(company)}* ({len(items)})"
        bullets = [_bullet(it) for it in items]
        blocks.extend(
            slack_client._section_with_header_and_bullets(header, bullets)
        )

    while len(blocks) > slack_client.MAX_BLOCKS and blocks:
        blocks.pop()
    return blocks


# --- Slack posting ------------------------------------------------------

async def _filter_invalid_urls_async(
    result: SectorResult, *, skip: bool,
) -> tuple[SectorResult, int]:
    if skip or not result.flat:
        return result, 0
    sem = asyncio.Semaphore(slack_client.URL_VALIDATION_CONCURRENCY)
    async with httpx.AsyncClient(
        timeout=config.URL_VALIDATION_TIMEOUT_S,
        headers={"User-Agent": "SignalAgent/0.1"},
        follow_redirects=True,
    ) as http:
        async def check(it: ImpactItem) -> tuple[str, bool]:
            async with sem:
                ok = await slack_client._validate_url_async(
                    it.story.canonical_url, http,
                )
                return it.story.id, ok
        results = await asyncio.gather(*(check(it) for it in result.flat))
    ok_ids = {sid for sid, ok in results if ok}
    dropped = sum(1 for _, ok in results if not ok)
    return _filter_result(result, ok_ids), dropped


def _filter_result(result: SectorResult, ok_ids: set[str]) -> SectorResult:
    grouped = {
        c: [it for it in items if it.story.id in ok_ids]
        for c, items in result.grouped.items()
    }
    grouped = {c: items for c, items in grouped.items() if items}
    flat = tuple(it for items in grouped.values() for it in items)
    return SectorResult(
        grouped=grouped,
        candidates_count=result.candidates_count,
        used_fallback=result.used_fallback,
        cost_usd=result.cost_usd,
        elapsed_seconds=result.elapsed_seconds,
        flat=flat,
    )


def post_sector_digest(
    result: SectorResult,
    *,
    digest_date: str,
    channel_id: str | None = None,
    channel_label: str | None = None,
    http: httpx.Client | None = None,
    skip_url_validation: bool = False,
    test_mode: bool = False,
) -> SlackResult:
    """Validate URLs, then post the company-grouped blocks to the sector channel.
    Reuses slack_client's transport (chat.postMessage / webhook) verbatim.
    ponytail: reuse slack transport internals rather than duplicate HTTP."""
    start = time.monotonic()
    channel_label = channel_label or config.SLACK_CHANNEL_LABEL_SECTOR

    bt = config.SLACK_BOT_TOKEN
    ch = channel_id if channel_id is not None else config.SLACK_CHANNEL_ID_SECTOR
    wh = config.SLACK_WEBHOOK_URL
    use_api = bool(bt and ch)
    if not use_api and not wh:
        return SlackResult(
            sent=False, channel_label=channel_label,
            stories_sent=0, stories_dropped_invalid_url=0,
            elapsed_seconds=round(time.monotonic() - start, 3),
            error="No Slack transport configured (need SLACK_BOT_TOKEN+"
                  "SLACK_CHANNEL_ID_SECTOR or SLACK_WEBHOOK_URL)",
        )

    own_http = http is None
    h = http or slack_client._make_default_http()
    blocks: list[dict] = []
    sent_ok = False
    error: str | None = None
    status: int | None = None
    slack_ts: str | None = None
    slack_channel: str | None = None
    dropped = 0
    try:
        if http is None:
            filtered, dropped = asyncio.run(
                _filter_invalid_urls_async(result, skip=skip_url_validation)
            )
        else:
            # Sync path (tests inject an http client): validate serially.
            ok_ids: set[str] = set()
            if skip_url_validation:
                ok_ids = {it.story.id for it in result.flat}
            else:
                for it in result.flat:
                    if slack_client.validate_url(it.story.canonical_url, http=h):
                        ok_ids.add(it.story.id)
                    else:
                        dropped += 1
            filtered = _filter_result(result, ok_ids)
        stories_sent = len(filtered.flat)
        blocks = build_sector_blocks(
            filtered, digest_date=digest_date, test_mode=test_mode,
        )
        text_prefix = "[TEST] " if test_mode else ""
        text = f"{text_prefix}Sector Signal — portfolio watch — {digest_date}"
        if use_api:
            sent_ok, status, error, slack_ts, slack_channel = \
                slack_client._post_via_api(
                    h=h, bot_token=bt, channel_id=ch, text=text, blocks=blocks,
                )
        else:
            sent_ok, status, error = slack_client._post_via_webhook(
                h=h, url=wh, text=text, blocks=blocks,
            )
    finally:
        if own_http:
            h.close()

    elapsed = round(time.monotonic() - start, 3)
    _log({
        "step": "sector_slack",
        "digest_date": digest_date,
        "sent": sent_ok,
        "channel_label": channel_label,
        "transport": "chat.postMessage" if use_api else "webhook",
        "stories_sent": stories_sent,
        "stories_dropped_invalid_url": dropped,
        "status_code": status,
        "error": error,
        "elapsed_seconds": elapsed,
    })
    return SlackResult(
        sent=sent_ok,
        channel_label=channel_label,
        stories_sent=stories_sent,
        stories_dropped_invalid_url=dropped,
        elapsed_seconds=elapsed,
        error=error,
        blocks=blocks,
        status_code=status,
        slack_ts=slack_ts,
        slack_channel=slack_channel,
    )


# --- Logging ------------------------------------------------------------

def _log_path() -> Path:
    day = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return config.LOGS_DIR / f"sector_{day}.jsonl"


def _log(rec: dict) -> None:
    rec.setdefault("ts", datetime.now(timezone.utc).isoformat(timespec="milliseconds"))
    with _log_path().open("a", encoding="utf-8") as f:
        f.write(json.dumps(rec, default=str) + "\n")
