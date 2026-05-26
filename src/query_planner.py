"""Layer 1: deterministic query planner.

Reads inputs/keywords.xlsx (single `Master Keywords` tab with a per-row Geo
column) and inputs/voices.xlsx (Tier column on India/US Top Voices, plus the
`New Additions` tab listing healthcare-active PE/VC firms). Emits a daily set
of QueryPlans split into four tracks:

  Track A — 13 priority plans (one per PriorityBucket × geo, AI gets +1 for
  the ventures/clinical split). These run every day and pull all keywords from
  the relevant sub-buckets, not 3 samples.

  Track B — ~15 rotation plans covering the long tail of non-priority
  sub-buckets. Deterministic 14-day cycle keyed on the digest date.

  Voice — 2 plans (India + US) naming Tier-1 voices from voices.xlsx.

  Firm — 1 plan naming PE/VC firms from the `New Additions` tab (PE/Strategics
  coverage anchor — distinct from the bucket-keyword Track A plan).

~31 plans/day total. Pure functions; same Excel + same date → same plans.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date as date_cls
from typing import Literal

from openpyxl import load_workbook

import config

Geography = Literal["India", "US", "Both"]
PlanGeo = Literal["India", "US", "Global"]

# Track B trims keyword samples per plan to keep prompts tight. Track A always
# uses every keyword in the target sub-buckets (richer prompt for priorities).
TRACK_B_KEYWORDS_PER_PLAN = 8

_GEO_LABEL: dict[PlanGeo, str] = {
    "India": "India",
    "US": "the United States",
    "Global": "globally (India, US, and cross-cutting)",
}


# --- Dataclasses --------------------------------------------------------

@dataclass(frozen=True)
class KeywordRow:
    bucket: str
    sub_bucket: str
    keyword: str
    geography: Geography   # "India", "US", or "Both"


@dataclass(frozen=True)
class Voice:
    """An individual healthcare voice. `tier` comes from column I of the India/US
    Top Voices tabs — 1 means "name this person inside the voice-anchored
    Perplexity prompt"; blank/anything else means "not tier-1"."""
    tier: int | None
    name: str
    category: str
    sub_domain: str
    role: str
    why: str
    reach_indicator: str
    linkedin_url: str
    geography: Literal["India", "US"]


@dataclass(frozen=True)
class Newsletter:
    name: str
    geography: str
    type_: str
    author: str
    description: str
    reach: str
    url: str


@dataclass(frozen=True)
class CompanyPage:
    """A row from the `Firms & Org Pages` tab — organizations to follow on
    LinkedIn (portfolio updates, etc.)."""
    geography: str
    name: str
    type_: str
    description: str
    why_follow: str
    linkedin_url: str


@dataclass(frozen=True)
class FirmAddition:
    """A row from the `New Additions` tab — PE/VC firms active in healthcare.
    The firm-anchored Perplexity query names these by firm + thesis so the
    fetcher can surface deal/portfolio news that bucket-keyword queries miss."""
    category: str          # "A — Local Indian PE/VC" etc.
    firm: str
    hq: str
    stage_ticket: str
    healthcare_thesis: str
    notable_portfolio: str
    source: str


@dataclass(frozen=True)
class QueryPlan:
    id: str
    geography: PlanGeo
    bucket: str
    sub_buckets: tuple[str, ...]
    keyword_sample: tuple[str, ...]
    keyword_count_total: int
    prompt_text: str
    track: Literal["A", "B", "voice", "firm"]
    priority_bucket: str | None = None     # PriorityBucket.key, or None for B/voice/firm
    voice_names: tuple[str, ...] = ()      # non-empty only for voice plans
    firm_names: tuple[str, ...] = ()       # non-empty only for firm plans


# --- Helpers ------------------------------------------------------------

def _slug(s: str) -> str:
    out: list[str] = []
    for ch in s.lower():
        if ch.isalnum():
            out.append(ch)
        elif out and out[-1] != "_":
            out.append("_")
    return "".join(out).strip("_")


def _to_int_or_none(v: object) -> int | None:
    if v is None:
        return None
    try:
        return int(str(v).strip())
    except (ValueError, TypeError):
        return None


def _s(v: object) -> str:
    return "" if v is None else str(v).strip()


def _is_blank(row: tuple) -> bool:
    return not row or not any(c is not None and _s(c) for c in row)


def _norm_geo(v: object) -> Geography:
    """Coerce the Excel Geo cell ('India', 'US', 'Both', or anything else) to a
    Geography literal. Unknown values default to 'Both' so a stray cell doesn't
    silently drop the keyword from every plan."""
    g = _s(v)
    if g.lower() == "india":
        return "India"
    if g.lower() == "us":
        return "US"
    return "Both"


def _geo_matches(row_geo: Geography, plan_geo: PlanGeo) -> bool:
    """Does this keyword row apply to a plan targeting plan_geo?

    - India plan accepts: India or Both
    - US plan accepts:    US or Both
    - Global plan accepts: anything
    """
    if plan_geo == "Global":
        return True
    if plan_geo == "India":
        return row_geo in ("India", "Both")
    if plan_geo == "US":
        return row_geo in ("US", "Both")
    return False


# --- Loaders ------------------------------------------------------------

def load_keywords() -> list[KeywordRow]:
    """Read the single `Master Keywords` tab. Columns: Bucket, Sub-bucket,
    Keyword, Geo. Header is row 1, data from row 2."""
    wb = load_workbook(config.KEYWORDS_XLSX, read_only=True, data_only=True)
    ws = wb["Master Keywords"]
    out: list[KeywordRow] = []
    for r in ws.iter_rows(values_only=True, min_row=2):
        if _is_blank(r):
            continue
        bucket, sub, kw = _s(r[0]), _s(r[1]), _s(r[2])
        if not kw:
            continue
        geo = _norm_geo(r[3] if len(r) > 3 else None)
        out.append(KeywordRow(bucket, sub, kw, geo))
    return out


def load_voices() -> list[Voice]:
    """Read India Top Voices + US Top Voices tabs. Layout:
       row 1 = banner (merged A1:H1), row 2 = blank, row 3 = header, row 4+ = data.
       The Tier column lives at column I (added 2026-05-26)."""
    wb = load_workbook(config.VOICES_XLSX, read_only=True, data_only=True)
    out: list[Voice] = []
    for tab, geo in (("India Top Voices", "India"), ("US Top Voices", "US")):
        ws = wb[tab]
        rows = list(ws.iter_rows(values_only=True))
        for r in rows[3:]:
            if _is_blank(r):
                continue
            cells = list(r) + [None] * max(0, 9 - len(r))
            name = _s(cells[1])
            if not name:
                continue
            out.append(Voice(
                tier=_to_int_or_none(cells[8]),
                name=name,
                category=_s(cells[2]),
                sub_domain=_s(cells[3]),
                role=_s(cells[4]),
                why=_s(cells[5]),
                reach_indicator=_s(cells[6]),
                linkedin_url=_s(cells[7]),
                geography=geo,
            ))
    return out


def load_newsletters() -> list[Newsletter]:
    """Read `Newsletters & Publications`. Layout: row 1 blank, row 2 header,
    row 3+ data. Columns: #, Publication, Geography, Type, Run by, What it
    covers, Reach, URL."""
    wb = load_workbook(config.VOICES_XLSX, read_only=True, data_only=True)
    ws = wb["Newsletters & Publications"]
    rows = list(ws.iter_rows(values_only=True))
    out: list[Newsletter] = []
    for r in rows[2:]:
        if _is_blank(r):
            continue
        cells = list(r) + [None] * max(0, 8 - len(r))
        name = _s(cells[1])
        if not name:
            continue
        out.append(Newsletter(
            name=name,
            geography=_s(cells[2]),
            type_=_s(cells[3]),
            author=_s(cells[4]),
            description=_s(cells[5]),
            reach=_s(cells[6]),
            url=_s(cells[7]),
        ))
    return out


def load_company_pages() -> list[CompanyPage]:
    """Read `Firms & Org Pages` (renamed from `Company Pages`). Layout: row 1
    blank, row 2 header, row 3+ data."""
    wb = load_workbook(config.VOICES_XLSX, read_only=True, data_only=True)
    ws = wb["Firms & Org Pages"]
    rows = list(ws.iter_rows(values_only=True))
    out: list[CompanyPage] = []
    for r in rows[2:]:
        if _is_blank(r):
            continue
        cells = list(r) + [None] * max(0, 7 - len(r))
        name = _s(cells[1])
        if not name:
            continue
        out.append(CompanyPage(
            geography=_s(cells[2]),
            name=name,
            type_=_s(cells[3]),
            description=_s(cells[4]),
            why_follow=_s(cells[5]),
            linkedin_url=_s(cells[6]),
        ))
    return out


def load_firm_additions() -> list[FirmAddition]:
    """Read the `New Additions` tab — PE/VC firms active in healthcare.
    Layout: row 1 blank, row 2 header, row 3+ data."""
    wb = load_workbook(config.VOICES_XLSX, read_only=True, data_only=True)
    if "New Additions" not in wb.sheetnames:
        return []
    ws = wb["New Additions"]
    rows = list(ws.iter_rows(values_only=True))
    out: list[FirmAddition] = []
    for r in rows[2:]:
        if _is_blank(r):
            continue
        cells = list(r) + [None] * max(0, 7 - len(r))
        firm = _s(cells[1])
        if not firm:
            continue
        out.append(FirmAddition(
            category=_s(cells[0]),
            firm=firm,
            hq=_s(cells[2]),
            stage_ticket=_s(cells[3]),
            healthcare_thesis=_s(cells[4]),
            notable_portfolio=_s(cells[5]),
            source=_s(cells[6]),
        ))
    return out


# --- Prompt templates ---------------------------------------------------

_TRACK_A_PROMPT_TEMPLATE = (
    "Healthcare news from the last 24 hours from {geo_label}, focused on {bucket_label}.\n"
    "Keywords to anchor on: {keyword_sample}.\n"
    "Prioritize: funding rounds >=$10M, regulatory actions (FDA/CDSCO/EMA), "
    "M&A >$100M, IPO filings, Phase 3 trial readouts, leadership moves at top-50 "
    "players, and substantive policy news. Skip listicles, opinion pieces, "
    "single-clinic openings, and routine PR.\n"
    "Return ONLY a JSON object (no markdown fences, no preamble):\n"
    "{{\n"
    '  "stories": [\n'
    '    {{"title": "headline", "url": "source URL", '
    '"published": "ISO 8601 datetime or null", "summary": "2-sentence summary"}}\n'
    "  ]\n"
    "}}"
)

_AI_VENTURES_PROMPT = (
    "AI in healthcare news from the last 24 hours — focus on NEW VENTURES: "
    "company launches, fundraising rounds (Series A through D), strategic "
    "partnerships, acquisitions, and product launches involving AI-native "
    "healthcare companies (e.g. Abridge, Hippocratic AI, OpenEvidence, "
    "Tennr, Innovaccer, Suki, Notable, Augmedix). Cover India and US.\n"
    "Skip: opinion pieces, listicles, conference recaps without news, "
    "incremental feature announcements.\n"
    "Return ONLY a JSON object (no markdown fences, no preamble):\n"
    "{{\n"
    '  "stories": [\n'
    '    {{"title": "headline", "url": "source URL", '
    '"published": "ISO 8601 datetime or null", "summary": "2-sentence summary"}}\n'
    "  ]\n"
    "}}"
)

_AI_CLINICAL_PROMPT = (
    "AI in healthcare news from the last 24 hours — focus on CLINICAL "
    "DEPLOYMENTS, REGULATION, and BIG PARTNERSHIPS: FDA clearances for AI "
    "devices, hospital-system AI rollouts, payer/provider partnerships with "
    "AI vendors, AI governance/regulation news (e.g. CMS AI rules, FDA AI "
    "framework), and large enterprise deals (Mayo, Cleveland Clinic, HCA, "
    "Apollo, Manipal). Cover India and US.\n"
    "Skip: opinion pieces, vendor blog posts, feature press releases.\n"
    "Return ONLY a JSON object (no markdown fences, no preamble):\n"
    "{{\n"
    '  "stories": [\n'
    '    {{"title": "headline", "url": "source URL", '
    '"published": "ISO 8601 datetime or null", "summary": "2-sentence summary"}}\n'
    "  ]\n"
    "}}"
)

_TRACK_B_PROMPT_TEMPLATE = (
    "Healthcare news from the last 24 hours from {geo_label}, focused on "
    "{bucket} / {sub_bucket}.\n"
    "Representative keywords: {keyword_sample}.\n"
    "Prioritize substantive announcements (funding, regulatory, M&A, leadership, "
    "product launches). Skip listicles and opinion pieces.\n"
    "Return ONLY a JSON object (no markdown fences, no preamble):\n"
    "{{\n"
    '  "stories": [\n'
    '    {{"title": "headline", "url": "source URL", '
    '"published": "ISO 8601 datetime or null", "summary": "2-sentence summary"}}\n'
    "  ]\n"
    "}}"
)

_VOICE_PROMPT_TEMPLATE = (
    "Substantive posts, articles, podcasts, or interviews published in the last 24 hours by "
    "these Tier-1 healthcare voices in {geo_label}: {voices}.\n"
    "Look across LinkedIn, X, Substacks, podcasts, op-eds, and media columns where applicable.\n"
    "Skip routine reposts, brief congratulatory replies, and short reactions. Focus on original "
    "takes, announcements, and substantive analysis.\n"
    "Return ONLY a JSON object (no markdown fences, no preamble):\n"
    "{{\n"
    '  "stories": [\n'
    '    {{"title": "voice name + topic", "url": "source URL", '
    '"published": "ISO 8601 datetime or null", "summary": "2-sentence summary"}}\n'
    "  ]\n"
    "}}"
)

_FIRM_PROMPT_TEMPLATE = (
    "Deal news, portfolio announcements, or fund actions from the last 24 hours involving "
    "these PE/VC firms active in Indian healthcare: {firms}.\n"
    "Surface: new investments, exits, portfolio company news (funding rounds, M&A, IPO), "
    "fund launches, leadership moves at the firms. Skip generic market commentary not tied "
    "to a named firm above.\n"
    "Return ONLY a JSON object (no markdown fences, no preamble):\n"
    "{{\n"
    '  "stories": [\n'
    '    {{"title": "headline", "url": "source URL", '
    '"published": "ISO 8601 datetime or null", "summary": "2-sentence summary"}}\n'
    "  ]\n"
    "}}"
)


# --- Track A: priority plans -------------------------------------------

def _bucket_label(bucket: config.PriorityBucket, plan_geo: PlanGeo) -> str:
    """Human-readable label for the prompt (e.g. 'Venture & IPO (India focus)')."""
    if plan_geo == "Global":
        return bucket.display
    return f"{bucket.display} ({plan_geo} focus)"


def _track_a_keywords_for(
    bucket: config.PriorityBucket, plan_geo: PlanGeo, kws: list[KeywordRow],
) -> list[str]:
    """All keywords from `bucket.sub_buckets` matching the plan's geo filter,
    preserving Excel order."""
    out: list[str] = []
    targets = set(bucket.sub_buckets)
    for kr in kws:
        if kr.sub_bucket not in targets:
            continue
        if not _geo_matches(kr.geography, plan_geo):
            continue
        out.append(kr.keyword)
    return out


def _build_track_a_plans(kws: list[KeywordRow]) -> list[QueryPlan]:
    plans: list[QueryPlan] = []
    for bucket in config.PRIORITY_BUCKETS:
        if bucket.key == "ai_healthcare":
            # AI gets two specialised plans, not the generic bucket sweep.
            plans.extend(_build_ai_plans(kws, bucket))
            continue
        for geo in bucket.geos:
            geo_lit: PlanGeo = geo  # narrowed by config typing
            keywords = _track_a_keywords_for(bucket, geo_lit, kws)
            prompt = _TRACK_A_PROMPT_TEMPLATE.format(
                geo_label=_GEO_LABEL[geo_lit],
                bucket_label=_bucket_label(bucket, geo_lit),
                keyword_sample=", ".join(keywords) if keywords else "(no keywords)",
            )
            plans.append(QueryPlan(
                id=f"pri__{bucket.key}__{geo_lit.lower()}",
                geography=geo_lit,
                bucket=bucket.display,
                sub_buckets=bucket.sub_buckets,
                keyword_sample=tuple(keywords),
                keyword_count_total=len(keywords),
                prompt_text=prompt,
                track="A",
                priority_bucket=bucket.key,
            ))
    return plans


def _build_ai_plans(
    kws: list[KeywordRow], bucket: config.PriorityBucket,
) -> list[QueryPlan]:
    """AI in Healthcare emits two plans — ventures/fundraises and clinical/
    partnerships — both tagged with priority_bucket='ai_healthcare' for Slack
    grouping. Keyword sample is shared (all sub-bucket keywords)."""
    keywords = _track_a_keywords_for(bucket, "Global", kws)
    out: list[QueryPlan] = []
    for variant_key, prompt_text in (
        ("ventures", _AI_VENTURES_PROMPT),
        ("clinical", _AI_CLINICAL_PROMPT),
    ):
        out.append(QueryPlan(
            id=f"pri__ai_healthcare__{variant_key}",
            geography="Global",
            bucket=bucket.display,
            sub_buckets=bucket.sub_buckets,
            keyword_sample=tuple(keywords),
            keyword_count_total=len(keywords),
            prompt_text=prompt_text,
            track="A",
            priority_bucket=bucket.key,
        ))
    return out


# --- Track B: rotated coverage of non-priority sub-buckets -------------

def _priority_sub_bucket_names() -> set[str]:
    out: set[str] = set()
    for b in config.PRIORITY_BUCKETS:
        out.update(b.sub_buckets)
    return out


def _all_non_priority_subs(kws: list[KeywordRow]) -> list[tuple[str, str, PlanGeo]]:
    """Return every (bucket, sub_bucket, geo) combination that has at least
    one keyword AND isn't already covered by a Track A priority sub-bucket.
    Geo per row determines the plan target: rows with Geo='India' → India plan,
    'US' → US plan, 'Both' → contribute to BOTH India and US plans for that sub.

    Deterministic ordering: bucket name asc, sub_bucket asc, geo (India, US).
    """
    priority_subs = _priority_sub_bucket_names()
    # Collect (bucket, sub) → set of geos that need plans
    needs: dict[tuple[str, str], set[PlanGeo]] = {}
    for kr in kws:
        if kr.sub_bucket in priority_subs:
            continue
        key = (kr.bucket, kr.sub_bucket)
        geos = needs.setdefault(key, set())
        if kr.geography in ("India", "Both"):
            geos.add("India")
        if kr.geography in ("US", "Both"):
            geos.add("US")

    out: list[tuple[str, str, PlanGeo]] = []
    for (b, s), gs in sorted(needs.items()):
        for g in ("India", "US"):
            if g in gs:
                out.append((b, s, g))   # type: ignore[arg-type]
    return out


def _rotation_pick(
    all_subs: list[tuple[str, str, PlanGeo]],
    today: date_cls,
    n: int,
) -> list[tuple[str, str, PlanGeo]]:
    """Pick `n` subs for today from the rotation. Deterministic: same date
    always yields the same pick. Strategy: pure contiguous-window math —
    `start = (ordinal * n) % len(all_subs)`, then take n consecutive subs with
    wrap-around. Consecutive days don't overlap; full universe is covered every
    `ceil(len(all_subs) / n)` days."""
    if not all_subs:
        return []
    start = (today.toordinal() * n) % len(all_subs)
    return [all_subs[(start + i) % len(all_subs)] for i in range(n)]


def _track_b_keywords_for(
    bucket: str, sub_bucket: str, plan_geo: PlanGeo, kws: list[KeywordRow],
) -> list[str]:
    out: list[str] = []
    for kr in kws:
        if kr.bucket != bucket or kr.sub_bucket != sub_bucket:
            continue
        if not _geo_matches(kr.geography, plan_geo):
            continue
        out.append(kr.keyword)
    return out


def _build_track_b_plans(
    kws: list[KeywordRow],
    today: date_cls,
) -> list[QueryPlan]:
    all_subs = _all_non_priority_subs(kws)
    picks = _rotation_pick(all_subs, today, config.TRACK_B_PLANS_PER_DAY)
    plans: list[QueryPlan] = []
    for bucket, sub_bucket, geo in picks:
        keywords = _track_b_keywords_for(bucket, sub_bucket, geo, kws)
        sample = keywords[:TRACK_B_KEYWORDS_PER_PLAN]
        prompt = _TRACK_B_PROMPT_TEMPLATE.format(
            geo_label=_GEO_LABEL[geo],
            bucket=bucket,
            sub_bucket=sub_bucket,
            keyword_sample=", ".join(sample),
        )
        plans.append(QueryPlan(
            id=f"tb__{_slug(sub_bucket)}__{geo.lower()}",
            geography=geo,
            bucket=bucket,
            sub_buckets=(sub_bucket,),
            keyword_sample=tuple(sample),
            keyword_count_total=len(keywords),
            prompt_text=prompt,
            track="B",
            priority_bucket=None,
        ))
    return plans


# --- Voice + Firm plans ------------------------------------------------

def _build_voice_plans() -> list[QueryPlan]:
    """One Tier-1 voice-anchored Perplexity query per geography. Reads the Tier
    column from voices.xlsx and names every Tier-1 voice in the prompt."""
    voices = load_voices()
    out: list[QueryPlan] = []
    for geo in ("India", "US"):
        tier1 = [v for v in voices if v.tier == 1 and v.geography == geo]
        if not tier1:
            continue
        names = tuple(v.name for v in tier1)
        prompt = _VOICE_PROMPT_TEMPLATE.format(
            geo_label=_GEO_LABEL[geo],
            voices=", ".join(names),
        )
        out.append(QueryPlan(
            id=f"voice__{geo.lower()}_t1",
            geography=geo,
            bucket="Tier-1 voices",
            sub_buckets=(),
            keyword_sample=(),
            keyword_count_total=0,
            prompt_text=prompt,
            track="voice",
            priority_bucket=None,
            voice_names=names,
        ))
    return out


def _build_firm_plans() -> list[QueryPlan]:
    """PE/VC firm-anchored query — names all firms from the `New Additions` tab
    so Perplexity surfaces deal/portfolio news the bucket-keyword PE plan would
    miss. Currently India-focused (the tab is India-heavy); US PE coverage rides
    on the Track A 'PE & Strategics' US plan."""
    firms = load_firm_additions()
    if not firms:
        return []
    firm_names = tuple(f.firm for f in firms)
    prompt = _FIRM_PROMPT_TEMPLATE.format(firms=", ".join(firm_names))
    return [QueryPlan(
        id="firm__india_pe_vc",
        geography="India",
        bucket="PE/VC firms (New Additions)",
        sub_buckets=(),
        keyword_sample=(),
        keyword_count_total=0,
        prompt_text=prompt,
        track="firm",
        priority_bucket=None,
        firm_names=firm_names,
    )]


# --- Orchestrator ------------------------------------------------------

def build_query_plans(*, today: date_cls | None = None) -> list[QueryPlan]:
    """Compose Track A + Track B + Voice + Firm plans for the given date.

    `today` defaults to UTC today — pass a fixed date for reproducible tests.
    """
    today = today or date_cls.today()
    kws = load_keywords()
    plans: list[QueryPlan] = []
    plans.extend(_build_track_a_plans(kws))
    plans.extend(_build_track_b_plans(kws, today))
    plans.extend(_build_voice_plans())
    plans.extend(_build_firm_plans())
    return plans


if __name__ == "__main__":
    plans = build_query_plans()
    print(f"Built {len(plans)} query plans.\n")
    by_track: dict[str, list[QueryPlan]] = {}
    for p in plans:
        by_track.setdefault(p.track, []).append(p)
    for track in ("A", "B", "voice", "firm"):
        bucket = by_track.get(track, [])
        print(f"  Track {track}: {len(bucket)} plans")
        for p in bucket:
            print(
                f"    [{p.geography:>6}] {p.id:<40} "
                f"kws={p.keyword_count_total:>3}  pri={p.priority_bucket or '-'}"
            )
    print("\n--- Example Track A prompt ---")
    print(by_track["A"][0].prompt_text)
